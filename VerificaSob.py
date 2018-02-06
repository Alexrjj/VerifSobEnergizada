from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import openpyxl
import os

#  Acessa os dados de login fora do script, salvo numa planilha existente, para proteger as informações de credenciais
dados = openpyxl.load_workbook('C:\\gomnet.xlsx')
login = dados['Plan1']
url = 'http://gomnet.ampla.com/'
username = login['C1'].value
password = login['C2'].value
wb = openpyxl.load_workbook('sobs.xlsx')

# ----------------- Modo Headless ------------------------
chromeOptions = webdriver.ChromeOptions()
prefs = {"download.default_directory" : os.getcwd(),
         "download.prompt_for_download": False}
chromeOptions.add_experimental_option("prefs",prefs)
chromeOptions.add_argument('--headless')
chromeOptions.add_argument('--window-size= 1600x900')
driver = webdriver.Chrome(chrome_options=chromeOptions)
# --------------------------------------------------------
# driver = webdriver.Chrome()

if __name__ == '__main__':
    driver.get(url)
    # Insere usuário e senha
    uname = driver.find_element_by_name('txtBoxLogin')
    uname.send_keys(username)
    passw = driver.find_element_by_name('txtBoxSenha')
    passw.send_keys(password)
    driver.find_element_by_id('ImageButton_Login').click()

    # Identifica o perfil "EMPREITEIRA-GESTOR" e loga no sistema
    perfil = Select(driver.find_element_by_id('ListBox_Perfil'))
    perfil.select_by_visible_text('EMPREITEIRA-GESTOR')
    driver.find_element_by_id('ImageButton_Logar').click()

    # Verifica se está energizada com base no número da SOB + número trabalho
    for sheet in wb.worksheets:
        for (trabalho, sob) in zip(sheet.iter_cols(min_row=2, min_col=1, max_col=1), sheet.iter_cols(min_row=2, min_col=2, max_col=2)):
            for (cell, cell2) in zip(trabalho, sob):
                driver.get('http://gomnet.ampla.com/DetalhesFiscalizacao.aspx?trabalho={}&OS={}'.format(cell.value, cell2.value))
                #Verifica se a sob está apta a fiscalizar (energizada) e retorna no console o resultado
                try:
                    energ = driver.find_element_by_xpath('//*[@id="Label_Data_Energizacao" and text() != ""]')
                    if energ.is_displayed():
                        log = open("status.txt", "a")
                        log.write(cell2.value + ' energizada' + '\n')
                        log.close()
                except NoSuchElementException:
                        log1 = open("status.txt", "a")
                        log1.write(cell2.value + ' não energizada' + '\n')
                        log1.close()
                        continue
